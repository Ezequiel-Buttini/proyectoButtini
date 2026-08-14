"""Tests for the 'Total' sheet reader: reads raw fields only, never trusts
the source file's own Consumo / Consumo Lts c/100km formula cells (those
get recomputed in the domain layer instead).
"""

from datetime import datetime

import openpyxl
import pytest

from app.adapters.excel_reader import OpenpyxlFuelLoadReader

HEADERS = [
    "Fecha de carga",
    "Fecha puente",
    "Responsable",
    "Turno",
    "Serie",
    "Coche",
    "Litros",
    "UREA",
    "Kms Odometro ",  # real header has a trailing space
    "Kms GPS Carga anterior",
    "Precinto NUEVO",
    "Precinto ANTERIOR",
    "TIPO DE COMBUSTIBLE",
    "CONSUMO",
    "CONSUMO LTS C/100KM)",
]


def _row(
    fecha_carga="2026-08-01 04:04:23",
    fecha_puente=None,
    responsable="023 - MARTINEZ JOSE LUIS",
    turno="M",
    serie=1461,
    coche=76,
    litros=109,
    urea=None,
    kms_odometro=0,
    kms_gps=447,
    precinto_nuevo=0,
    precinto_anterior=0,
    tipo_combustible="INFINIA",
):
    return [
        fecha_carga,
        fecha_puente if fecha_puente is not None else datetime(2026, 8, 1),
        responsable,
        turno,
        serie,
        coche,
        litros,
        urea,
        kms_odometro,
        kms_gps,
        precinto_nuevo,
        precinto_anterior,
        tipo_combustible,
        999,  # CONSUMO in the source -- must be ignored, not read
        "should be ignored too",  # CONSUMO LTS C/100KM) -- must be ignored
    ]


def _build_workbook(tmp_path, rows, sheet_name="Total", extra_sheet_first=True):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if extra_sheet_first:
        wb.create_sheet("Parque Movil")  # the real workbook has 5 other sheets
    sheet = wb.create_sheet(sheet_name)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    path = tmp_path / "input_sample.xlsx"
    wb.save(path)
    return path


def test_reads_all_raw_fields_from_the_total_sheet(tmp_path):
    path = _build_workbook(tmp_path, [_row()])
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert len(records) == 1
    record = records[0]
    assert record.fecha_carga == datetime(2026, 8, 1, 4, 4, 23)
    assert record.fecha_puente == datetime(2026, 8, 1)
    assert record.responsable == "023 - MARTINEZ JOSE LUIS"
    assert record.turno == "M"
    assert record.serie == "1461"
    assert record.coche == "76"
    assert record.litros == 109
    assert record.urea is None
    assert record.kms_odometro == 0
    assert record.kms_gps_carga_anterior == 447
    assert record.precinto_nuevo == 0
    assert record.precinto_anterior == 0
    assert record.tipo_combustible == "INFINIA"


def test_keeps_non_numeric_serie_and_coche_as_text(tmp_path):
    path = _build_workbook(tmp_path, [_row(serie="TURISMO", coche="Taller")])
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert records[0].serie == "TURISMO"
    assert records[0].coche == "Taller"


def test_blank_serie_becomes_empty_text_not_literal_none(tmp_path):
    path = _build_workbook(tmp_path, [_row(serie=None, coche="Taller")])
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert records[0].serie == ""


def test_blank_tipo_combustible_becomes_empty_text(tmp_path):
    path = _build_workbook(tmp_path, [_row(tipo_combustible=None)])
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert records[0].tipo_combustible == ""


def test_formats_float_serie_without_spurious_decimal(tmp_path):
    path = _build_workbook(tmp_path, [_row(serie=400301.0)])
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert records[0].serie == "400301"


def test_ignores_extra_unnamed_trailing_columns(tmp_path):
    """Real file has one stray value in an unnamed 16th column -- must not
    break reading or leak into any field."""
    path = _build_workbook(tmp_path, [_row()])
    wb = openpyxl.load_workbook(path)
    wb["Total"].cell(row=2, column=16, value=421)
    wb.save(path)
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert len(records) == 1


def test_reads_the_total_sheet_by_name_not_by_position(tmp_path):
    path = _build_workbook(tmp_path, [_row()], extra_sheet_first=True)
    reader = OpenpyxlFuelLoadReader()

    records = reader.read(path)

    assert len(records) == 1

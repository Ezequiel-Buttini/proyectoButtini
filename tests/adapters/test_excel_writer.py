from datetime import datetime

import openpyxl

from app.adapters.excel_writer import OpenpyxlReportWriter
from app.domain.models import CleanedRecord, FuelLoadRecord

EXPECTED_HEADER = [
    "Fecha de carga",
    "Fecha puente",
    "Responsable",
    "Turno",
    "Serie",
    "Coche",
    "Litros",
    "UREA",
    "Kms Odometro",
    "Kms GPS Carga anterior",
    "Precinto Nuevo",
    "Precinto Anterior",
    "Tipo de Combustible",
    "Consumo",
    "Consumo Lts c/100km",
]


def _record(**overrides):
    defaults = dict(
        fecha_carga=datetime(2026, 8, 1, 4, 4, 23),
        fecha_puente=datetime(2026, 8, 1),
        responsable="023 - MARTINEZ JOSE LUIS",
        turno="M",
        serie="1461",
        coche="76",
        litros=109,
        urea=None,
        kms_odometro=0,
        kms_gps_carga_anterior=447,
        precinto_nuevo=0,
        precinto_anterior=0,
        tipo_combustible="INFINIA",
    )
    defaults.update(overrides)
    return FuelLoadRecord(**defaults)


def _cleaned_row(record=None, consumo=24.38, consumo_lts_c_100km="||" * 24):
    return CleanedRecord(
        record=record or _record(),
        consumo=consumo,
        consumo_lts_c_100km=consumo_lts_c_100km,
    )


def test_writes_header_and_one_row_per_record(tmp_path):
    path = tmp_path / "output.xlsx"

    OpenpyxlReportWriter().write([_cleaned_row()], path)

    wb = openpyxl.load_workbook(path)
    rows = list(wb.active.iter_rows(values_only=True))

    assert rows[0] == tuple(EXPECTED_HEADER)
    assert rows[1] == (
        "2026-08-01 04:04:23",
        datetime(2026, 8, 1),
        "023 - MARTINEZ JOSE LUIS",
        "M",
        "1461",
        "76",
        109,
        None,
        0,
        447,
        0,
        0,
        "INFINIA",
        24.38,
        "||" * 24,
    )


def test_writes_urea_value_when_present(tmp_path):
    path = tmp_path / "output.xlsx"

    OpenpyxlReportWriter().write([_cleaned_row(record=_record(urea=15))], path)

    wb = openpyxl.load_workbook(path)
    rows = list(wb.active.iter_rows(values_only=True))

    assert rows[1][7] == 15


def test_writes_no_total_rows_just_one_row_per_record(tmp_path):
    path = tmp_path / "output.xlsx"
    rows_in = [
        _cleaned_row(record=_record(coche="1")),
        _cleaned_row(record=_record(coche="2")),
    ]

    OpenpyxlReportWriter().write(rows_in, path)

    wb = openpyxl.load_workbook(path)
    rows = list(wb.active.iter_rows(values_only=True))

    # header + 2 data rows, nothing else
    assert len(rows) == 3
